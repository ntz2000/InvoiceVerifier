from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.models import CompareRow


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
MATCH_FILL = PatternFill("solid", fgColor="E2F0D9")
MISMATCH_FILL = PatternFill("solid", fgColor="F4CCCC")
UNCERTAIN_FILL = PatternFill("solid", fgColor="FFF2CC")


def _is_uncertain(row: CompareRow) -> bool:
    return any(keyword in row.message for keyword in ("缺失", "无数据", "无法判断", "复核", "不确定", "待核验", "人工核验"))


def _row_status_text(row: CompareRow) -> str:
    if row.is_match:
        return "一致"
    if _is_uncertain(row):
        return "无数据/待核验"
    return "不一致"


def generate_excel_report(rows: List[CompareRow], output_path: str, conclusion: str) -> None:
    """用 openpyxl 生成一致性比对报告。"""
    if not output_path:
        raise ValueError("报告输出路径不能为空。")

    target_path = Path(output_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "比对报告"

    sheet.merge_cells("A1:E1")
    title_cell = sheet["A1"]
    title_cell.value = "电子发票查验一致性比对报告"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.merge_cells("A2:E2")
    conclusion_cell = sheet["A2"]
    conclusion_cell.value = f"审核结论：{conclusion}"
    conclusion_cell.font = Font(size=12, bold=True)
    conclusion_cell.alignment = Alignment(horizontal="left", vertical="center")

    headers = ["字段名", "用户上传发票", "官方查验结果", "是否一致", "差异说明"]
    header_row = 4
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, compare_row in enumerate(rows, start=header_row + 1):
        values = [
            compare_row.field_name,
            compare_row.user_value,
            compare_row.official_value,
            _row_status_text(compare_row),
            compare_row.message,
        ]

        if compare_row.is_match:
            row_fill = MATCH_FILL
        elif _is_uncertain(compare_row):
            row_fill = UNCERTAIN_FILL
        else:
            row_fill = MISMATCH_FILL

        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 冻结标题、结论和表头，滚动大量字段时仍能看到表头。
    sheet.freeze_panes = "A5"

    for column_index in range(1, len(headers) + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0
        for cell in sheet[column_letter]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 4, 12), 48)

    workbook.save(str(target_path))
